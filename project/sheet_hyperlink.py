# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2020/03/26
# @Author  : 陈志鹏
# @File    : xxx.py

"""
为所有工作表生成超链接, 存为目录工作表
"""

import os
import project.log
import time
import pandas as pd
import openpyxl
from PySide2.QtCore import *

logger = project.log.Log(__name__).getlog()

nowTime = lambda:int(round(time.time()*1000))

class SheetHyperLink(QObject):
    #message_singel = Signal(str)
    finish_singel = Signal()
    statusBar_singel = Signal(str)

    def __init__(self):
        super(SheetHyperLink, self).__init__()

    def __del__(self):
        print('delete %s' % self.__class__.__name__)

    def set_parameter(self, *args, ** kwargs):
        self.dst_file = kwargs.get('dst_file')

    def open_dst_file(self):
        # self.writer = pd.ExcelWriter(self.dst_file, engine='openpyxl', data_only=True)
        self.writer = pd.ExcelWriter(self.dst_file, engine='openpyxl')
        # try to open an existing workbook
        # self.writer.book = openpyxl.load_workbook(self.dst_file, data_only=True)
        self.writer.book = openpyxl.load_workbook(self.dst_file)

        # copy existing sheets
        self.writer.sheets = {ws.title: ws for ws in self.writer.book.worksheets}

    def save_dst_file(self):
        self.writer.save()

    def modify_dst_file(self, dataframe, **to_excel_kwargs):
        dataframe.to_excel(self.writer, **to_excel_kwargs)

    def hyperlink_for_sheets(self):
        if '目录' in self.writer.book.sheetnames:
            wsheet = self.writer.book['目录']
            print('exist')
        else:
            wsheet = self.writer.book.create_sheet("目录")
            print('not exist')

        for i, sheet in enumerate(self.writer.sheets.keys()):
            link = '=HYPERLINK("#{}{}{}!A1","{}")'.format("'", sheet, "'", sheet)
            wsheet.cell(row=i+1, column=2, value=link)

    def run(self):
        self.open_dst_file()
        self.hyperlink_for_sheets()
        self.save_dst_file()


if __name__ == '__main__':
    starttime = nowTime()
    sheet_hyperlink = SheetHyperLink()
    sheet_hyperlink.set_parameter(dst_file=r'../datas/202004安徽天孚报表-费用.xlsx')

    sheet_hyperlink.run()
    # merge_expense.check()
    print('finish', nowTime()-starttime, 'ms')
