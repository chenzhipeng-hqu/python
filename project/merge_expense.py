# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2020/03/26
# @Author  : 陈志鹏
# @File    : xxx.py

"""
提取一个多个工作簿的sheet信息， 合并到费用报表底稿文件
"""

import os
import project.log
import time
import pandas as pd
import openpyxl
from PySide2.QtCore import *

logger = project.log.Log(__name__).getlog()

nowTime = lambda:int(round(time.time()*1000))

class MergeExpense(QObject):
    #message_singel = Signal(str)
    finish_singel = Signal()
    statusBar_singel = Signal(str)

    def __init__(self):
        super(MergeExpense, self).__init__()
        self.expense_table = [
            #0.identification, 1.expense, 2.copy_row, 3.read_row, 4.read_col, 5.other_cnt
            # ['G', '管理费用', 92-5, 4, [0, 1, 2]],
            # ['Y', '研发费用', 92-5, 4, [0, 1, 2]],
            # ['X', '营业费用', 92-5, 4, [0, 1, 2]],
            # ['X', '销售费用', 92-5, 4, [0, 1, 2]],
            # ['C', '财务费用', 12-5, 4, [0, 1]],
            # ['Z', '制造费用', 92-5, 4, [0, 1, 2]],
            ['G', '管理费用', 92-5, 4, [2], 5],
            ['Y', '研发费用', 92-5, 4, [2], 3],
            ['X', '营业费用', 92-5, 4, [2], 10],
            ['X', '销售费用', 92-5, 4, [2], 10],
            ['C', '财务费用', 12-5, 4, [1], 0],
            ['Z', '制造费用', 92-5, 4, [2], 4],
        ]

        self.company_table = [
            #0.company in file  #1.company in sheet  #2. handle_G  #3. handle_Y  #4. handle_X  #5. handle_X2  #6. handle_C  #6. handle_Z
            ['高新', '高新', self.handle_GYXZ_other_data, self.handle_GYXZ_other_data, self.handle_GYXZ_other_data, self.handle_GYXZ_other_data, self.handle_general, self.handle_Z_GX],
            ['有机硅', '有机硅', self.handle_general, None, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['香港', '香港', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['九江天祺', '天祺', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['九江天赐', '九江', self.handle_G_JJTC, self.handle_YZ_JJTC, self.handle_X_JJTC, self.handle_X_JJTC, self.handle_general, self.handle_YZ_JJTC],
            ['池州天赐', '池州天赐', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['江西创新', '创新中心', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['宜春天赐', '宜春天赐', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['中硝', '中硝', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['中天鸿锂', '中天鸿锂', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['天津', '天津', self.handle_GYXZ_other_data, self.handle_GYXZ_other_data, self.handle_X_TJ, self.handle_X_TJ, self.handle_general, self.handle_GYXZ_other_data],
            ['安徽天孚', '安徽天孚', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['宁德', '宁德', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['捷克', '捷克天赐', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
            ['浙江天硕', '浙江天硕', self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general, self.handle_general],
        ]

    def __del__(self):
        print('delete %s' % self.__class__.__name__)

    def set_parameter(self, *args, ** kwargs):
        self.month = kwargs.get('month')
        self.file_path = kwargs.get('file_path')
        self.dst_file = kwargs.get('dst_file')

    def open_dst_file(self):
        # self.writer = pd.ExcelWriter(self.dst_file, engine='openpyxl', data_only=True)
        self.writer = pd.ExcelWriter(self.dst_file, engine='openpyxl')
        # try to open an existing workbook
        # self.writer.book = openpyxl.load_workbook(self.dst_file, data_only=True)
        self.writer.book = openpyxl.load_workbook(self.dst_file)

        # copy existing sheets
        self.writer.sheets = {ws.title: ws for ws in self.writer.book.worksheets}

    def save_dst_file(self):
        self.writer.save()

    def modify_dst_file(self, dataframe, **to_excel_kwargs):
        dataframe.to_excel(self.writer, **to_excel_kwargs)

    def search_files(self, path):
        filelist = []
        for root, dirs, files in os.walk(path, topdown=False):
            for name in files:
                str = os.path.join(root, name)
                if str.endswith('.xlsx') or str.endswith('.xls'):
                    if str.endswith('merge.xlsx') or str.find('~') != -1:
                        continue
                    else:
                        filelist.append(str)
                        # break # just test
        return filelist

    def hyperlink_for_sheets(self):
        wsheet = self.writer.book['目录']
        for i, sheet in enumerate(self.writer.sheets.keys()):
            link = r'{}#{}{}{}!E1'.format(os.path.basename(self.dst_file), "'", sheet, "'")
            wsheet.cell(row=i+3, column=2, value=sheet).hyperlink = link

    def find_company_in_filename(self, src_file):
        for company in self.company_table:
            if company[0] in src_file:
                return company

    def find_src_sheet(self, match, sheets):
        for sheet in sheets:
            if match in sheet:
                return sheet
        return ''

    def find_dst_sheet(self, match, sheets):
        for sheet in sheets:
            if match == sheet:
                return sheet
        return ''

    def find_dst_col(self, match, sheet):
        for i in range(1, sheet.max_column+1):
            value = sheet.cell(row=5, column=i).value
            # print(type(match), ', ', match, ';  ', type(value), ', ', value)
            if str(match) in str(value):
                # print(value, i)
                # return i #just for test
                return i - 1
        return 0

    def handle_general(self, src_df, expense, dst_sheet, dst_month_col):
        src_data = src_df[self.month][:expense[2]]
        src_data.astype(dtype='float64')
        self.modify_dst_file(src_data, sheet_name=dst_sheet, startcol=dst_month_col, startrow=expense[3] + 1,
                             index=False, header=False)

    def handle_G_JJTC(self, src_df, expense, dst_sheet, dst_month_col):
        self.handle_general(src_df, expense, dst_sheet, dst_month_col)
        src_data_other = src_df[self.month][expense[2] + 1:expense[2] + 2]
        src_data_other2 = src_df[self.month][expense[2] + 3:expense[2] + 1 + 5]
        src_data_other3 = src_df[self.month][expense[2] + 7:expense[2] + 1 + 8]
        src_data_other = pd.concat([src_data_other, src_data_other2, src_data_other3])

        src_data_other.astype(dtype='float64')
        self.modify_dst_file(src_data_other, sheet_name=dst_sheet, startcol=dst_month_col,
                             startrow=expense[2] + 6, index=False, header=False)

    def handle_X_JJTC(self, src_df, expense, dst_sheet, dst_month_col):
        self.handle_general(src_df, expense, dst_sheet, dst_month_col)
        src_data_other = src_df[self.month][expense[2] + 1:expense[2] + 3]
        src_data_other2 = src_df[self.month][expense[2] + 3:expense[2] + 1 + 6]
        src_data_other3 = src_df[self.month][expense[2] + 3:expense[2] + 1 + 3]
        src_data_other4 = src_df[self.month][expense[2] + 6:expense[2] + 1 + 8]
        src_data_other = pd.concat([src_data_other, src_data_other3, src_data_other2, src_data_other4])

        src_data_other.astype(dtype='float64')
        self.modify_dst_file(src_data_other, sheet_name=dst_sheet, startcol=dst_month_col,
                             startrow=expense[2] + 6, index=False, header=False)

    def handle_YZ_JJTC(self, src_df, expense, dst_sheet, dst_month_col):
        self.handle_general(src_df, expense, dst_sheet, dst_month_col)
        src_data_other = src_df[self.month][expense[2] + 1:expense[2] + 2]
        src_data_other2 = src_df[self.month][expense[2] + 3:expense[2] + 1 + 5]
        src_data_other = pd.concat([src_data_other, src_data_other2])

        src_data_other.astype(dtype='float64')
        self.modify_dst_file(src_data_other, sheet_name=dst_sheet, startcol=dst_month_col,
                             startrow=expense[2] + 6, index=False, header=False)

    def handle_Z_GX(self, src_df, expense, dst_sheet, dst_month_col):
        self.handle_general(src_df, expense, dst_sheet, dst_month_col)
        src_data_other = src_df[self.month][expense[2] + 1:expense[2] + 1 + 3]
        src_data_other2 = src_df[self.month][expense[2] + 1 + 4:expense[2] + 1 + 5]
        src_data_other = pd.concat([src_data_other, src_data_other2])

        src_data_other.astype(dtype='float64')
        self.modify_dst_file(src_data_other, sheet_name=dst_sheet, startcol=dst_month_col,
                             startrow=expense[2] + 6, index=False, header=False)

    def handle_X_TJ(self, src_df, expense, dst_sheet, dst_month_col):
        self.handle_general(src_df, expense, dst_sheet, dst_month_col)
        src_data_other = src_df[self.month][expense[2] + 1:expense[2] + 1 + 6]
        src_data_other2 = src_df[self.month][expense[2] + 1 + 8:expense[2] + 1 + 12]
        src_data_other = pd.concat([src_data_other, src_data_other2])
        src_data_other.astype(dtype='float64')
        self.modify_dst_file(src_data_other, sheet_name=dst_sheet, startcol=dst_month_col,
                             startrow=expense[2] + 6, index=False, header=False)

    def handle_GYXZ_other_data(self, src_df, expense, dst_sheet, dst_month_col):
        self.handle_general(src_df, expense, dst_sheet, dst_month_col)
        src_data_other = src_df[self.month][expense[2] + 1:expense[2] + 1 + expense[5]]
        src_data_other.astype(dtype='float64')
        self.modify_dst_file(src_data_other, sheet_name=dst_sheet, startcol=dst_month_col,
                             startrow=expense[2] + 6, index=False, header=False)

    def run(self):
        """
        1. 打开文件夹里待提取的第一个文件
        2. 提取 管理费用(G), 研究开发费用(Y), 销售费用(X), 财务费用(C), 制造费用(Z), 数据
        3. 打开待合并文件， 填入相应位置
        """
        self.statusBar_singel.emit('正在合并...\r\n')
        filelist = self.search_files(self.file_path)
        # print(filelist)
        print(len(filelist))

        self.open_dst_file()
        print(self.writer.sheets.keys())

        self.hyperlink_for_sheets()

        for src_file in filelist:
            print(src_file)
            company = self.find_company_in_filename(src_file)
            excel_file = pd.ExcelFile(src_file)
            print(excel_file.sheet_names)
            for i, expense in enumerate(self.expense_table):
                self.statusBar_singel.emit('正在合并%s %s\r\n' % (os.path.basename(src_file), expense[1]))
                print(expense, end=', ')

                # 找到源文件对应的费用名称
                src_sheet_match = '2020实际' + expense[1]
                src_sheet = self.find_src_sheet(src_sheet_match, excel_file.sheet_names)
                print(src_sheet, end=', ')

                # 找到目标文件对应公司对应的费用名称费用的表格
                dst_sheet_match = expense[0]+'-'+company[1]
                dst_sheet = self.find_dst_sheet(dst_sheet_match, self.writer.sheets)
                print(dst_sheet, end=', ')

                if dst_sheet != '' and src_sheet != '' and company[i+2] != None:
                    wsheet = self.writer.book[dst_sheet]
                    month = '2020年' + self.month
                    dst_month_col = self.find_dst_col(self.month, wsheet)
                    assert(dst_month_col != 0)
                    print('[dst_col:%d]' % (dst_month_col), end=', ')

                    src_df = pd.read_excel(src_file, sheet_name=src_sheet, header=expense[3])
                    src_df_row_max =  len(src_df[self.month])
                    print('[src_data_len:%d]' % (src_df_row_max), end=', ')
                    assert (src_df_row_max > expense[2]), "行数太少"
                    # src_month_col = list(src_df.columns.values).index(self.month)
                    # print('[src_col:%d]' % (src_month_col), end=', ')

                    company[i+2](src_df, expense, dst_sheet, dst_month_col)

                print('')
            print('')
            excel_file.close()
        self.save_dst_file()
        self.statusBar_singel.emit('完成.\r\n')
        self.finish_singel.emit()

    def check(self):
        pass


if __name__ == '__main__':
    starttime = nowTime()
    merge_expense = MergeExpense()
    # merge_expense.set_parameter(month='4月', file_path=r'../datas/费用合并底稿/202004',
    #                         dst_file=r'../datas/费用合并底稿/202004费用合并（公式）底稿(副本).xlsm')
    merge_expense.set_parameter(month='5月', file_path=r'../datas/费用合并底稿/202005',
                            dst_file=r'../datas/费用合并底稿/202005费用合并（公式）底稿 - 副本.xlsm')

    merge_expense.run()
    # merge_expense.check()
    print('finish', nowTime()-starttime, 'ms')
