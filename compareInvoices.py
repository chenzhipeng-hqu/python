#! /usr/bin/python3
#  -*- coding:utf-8 -*-

import re
import os
import sys
import copy
import time
import random
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

__author__ = 'chenzhipeng'

"""
命令行解析器

Usage:
	./compareInvoices.py file_path
"""


class Invoice:
    '''
    找连号发票
    '''

    def __init__(self, file_name):
        self.file_name = file_name
        self.workBook = self._readExcel(file_name)
        self.invoices = self._getInvoices(self.workBook)
        self.fgColor_list = list()

    def creatExcel(self):
        pass

    def _readExcel(self, file_name):
        return openpyxl.load_workbook(file_name)

    def _getInvoices(self, workBook):
        self.sheet_names = workBook.sheetnames
        print (workBook.sheetnames) # [u'sheet1', u'sheet2']

        invoices = []
        for sheet in workBook:
            sheet_invoices = []
            # print(sheet.max_row, sheet.max_column)
            for row in range(3, sheet.max_row+1):
                row_invoices = []
                for col in range(1, sheet.max_column+1):
                    # temp = sheet.cell(2, col).value
                    # if temp != '汇总':
                        # print(temp)
                        # print(type(temp))
                    data = sheet.cell(row, col).value
                    if sheet.cell(2, col).value != '汇总' and re.match('^[0-9]+$', str(data)):
                    # if data != None:
                    # if str(data).isdigit():
                        row_invoices.append(int(data))
                    else:
                        row_invoices.append(int(0))

                sheet_invoices.append(row_invoices)
            invoices.append(sheet_invoices)

        # invoices = [[[int(sheet.cell(row, col).value) for col in range(1, sheet.max_column) if sheet.cell(row, col).value != None] for row in range(3, sheet.max_row)] for sheet in workBook]
        # a = 0x01 if 0x00 else 0x02
        # print(a)
        # print(invoices)
        return invoices

    def convertPositon(self):
        pass

    def signStyle(self):
        pass

    def findConsecutiveNum(self):
        invoices = self.invoices
        # print(invoices)
        # self._getIndex(11111170, self.invoices)
        position_set = list()
        position_without_color = list()

        fgColorOffset = str(0)
        for sheet_num, sheet_invoices in enumerate(invoices):
            for row, row_invoices in enumerate(sheet_invoices):
                for col, invoice in enumerate(row_invoices):
                    position_b = self._getIndex(invoice + 1, invoices)
                    if len(position_b) > 0:
                        # print('----------------------------start-----------------------------')
                        position_a = self._getIndex(invoice, invoices)

                        if (position_a[0] in position_without_color):
                            position_color = position_without_color.index(position_a[0])
                            # if position_set[position_color].__len__() < 4:

                                # print('%10s %2d%02s (%08d)' % (self.sheet_names[position_set[position_color][0]], position_set[position_color][1]+3, get_column_letter(position_set[position_color][2]+1), invoice+1))
                                # print(position_set)
                                # print(position_set[position_color])

                            fgColorOffset = position_set[position_color][3]
                        elif (position_b[0] in position_without_color):
                            position_color = position_without_color.index(position_b[0])
                            fgColorOffset = position_set[position_color][3]
                        else:
                            fgColorOffset = '{:06x}'.format(self._generateRand())

                        # for idx, ppp in enumerate(position_without_color):
                            # print(idx, ppp)

                        for idx, pos in enumerate(position_a):
                            position_without_color.append(position_a[idx])
                        for idx, pos in enumerate(position_b):
                            position_without_color.append(position_b[idx])

                        position = copy.deepcopy(position_a)
                        position1 = copy.deepcopy(position_b)

                        for idx, pos in enumerate(position):
                            position[idx].append(fgColorOffset)
                        for idx, pos in enumerate(position1):
                            position1[idx].append(fgColorOffset)

                        position_set += position
                        position_set += position1
                        # print(position_set)
                        print('%10s %2d%02s (%08d)' % (self.sheet_names[position[0][0]], position[0][1]+3, get_column_letter(position[0][2]+1), invoice), end=' 连号 ')
                        print('%10s %2d%02s (%08d)' % (self.sheet_names[position1[0][0]], position1[0][1]+3, get_column_letter(position1[0][2]+1), invoice+1))
                        # print('%10s %2d%02s (%08d) %6s' % (self.sheet_names[position[0][0]], position[0][1]+3, get_column_letter(position[0][2]+1), invoice, position[0][3]), end=' 连号 ')
                        # print('%10s %2d%02s (%08d) %6s' % (self.sheet_names[position1[0][0]], position1[0][1]+3, get_column_letter(position1[0][2]+1), invoice+1, position[0][3]))
                        # print('{} {}{}({}) {}'.format(self.sheet_names[position[0][0]], position[0][1], position[0][2], invoice, position[0][3]), end=' 连号 ')
                        # print('{} {}{}({}) {}'.format(self.sheet_names[position1[0][0]], position1[0][1], position1[0][2], invoice+1, position1[0][3]))
                        # print(position_without_color)
                        # print(id(position_without_color[0]), id(position_set[0]))
                        # print('            ----------------end-----------------------------')

        # print(position_set)
        position_=list(set([tuple(t) for t in position_set]))
        # print(position_)
        return position_

    def _generateRand(self):
        result = random.randrange(0x0000ff, 0xff0000, 0x0000ff)
        if result not in self.fgColor_list:
            self.fgColor_list.append(result)
        else:
            print('repetition.')
            self._generateRand()

        return result

    def signConsecutiveNum(self, signPositionList):
        # print(signPositionList)

        for sheet, row, col, fgColor in signPositionList:
            # print(sheet, row, col)
            # print(self.sheet_names[sheet])
            # print(type(self.sheet_names[sheet]))
            sheet = self.workBook[self.sheet_names[sheet]]
            # sheet = self.workBook.get_active_sheet()
            bold_itatic_24_font = openpyxl.styles.Font(color=openpyxl.styles.colors.WHITE)
            # bold_itatic_24_font = openpyxl.styles.Font(name='等线', size=24, italic=True, color=openpyxl.styles.colors.RED, bold=True)
            # position = str(str(get_column_letter(col))+str(row))
            # signPosition = sheet['F4']
            signPosition = sheet.cell(row=row+1+2, column=col+1)
            # signPosition = sheet[position]
            signPosition.font = bold_itatic_24_font
            # {'lightGrid', 'gray0625', 'lightTrellis', 'lightDown', 'lightVertical', 'darkTrellis', 'darkHorizontal', 'darkVertical', 'darkGrid', 'darkGray', 'solid', 'darkUp', 'lightGray', 'mediumGray', 'darkDown', 'lightHorizontal', 'lightUp', 'gray125'}
            # print('----------------------------------------------------------------------')
            # fgColor = 0xFFA500
            # print(type("FFA500"))
            # print(type(fgColor))
            # print(type(str(fgColor)+str(0xaa)))
            # print(str(fgColor)+str(0xaa))
            # print('----------------------------------------------------------------------')
            # print(str(hex(fgColor)))
            fill = openpyxl.styles.PatternFill('solid', fgColor=str(fgColor))
            signPosition.fill = fill

        file_name_result = os.path.splitext(os.path.basename(self.file_name))[0]+'_result.xlsx'
        self.workBook.save(file_name_result)
        print('Finished! Please open '+file_name_result+'.')
        pass

    def _getIndex(self, value, invoices):
        """获得二维列表某个值的一维索引值
            思想：先选出包含value值的一维列表，然后判断此一维列表在二维列表中的索引 
        """
        position = []
        for sheet_num, sheet_invoices in enumerate(invoices):
            for row, row_invoices in enumerate(sheet_invoices):
                for col, invoice in enumerate(row_invoices):
                    if value == invoice:
                        position.append([sheet_num, row, col])
        # print(position)
        return position

    def start(self):
        try:
            signPositionList = self.findConsecutiveNum()
            self.signConsecutiveNum(signPositionList)
        except Exception as err:
            print('ERROR: ', end='')
            print(err)
        pass


def printInfo():
    print('#============================================================')
    print('#  FileName    : '+os.path.basename(__file__)+'                            ')
    print('#  Author      : '+__author__+' ')
    print('#  Copyright   : Feather (c) 2018')
    print('#  Time-stamp  : < '+time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))+' >')
    print('#  Description : compareInvoices to find consecutive number.')
    print('#============================================================')
    print('')

if __name__ == '__main__':
    printInfo()

    if(len(sys.argv) > 1):
        print(sys.argv[1])
        print('')
        invoices = Invoice(sys.argv[1])
        invoices.start()
    else:
        print('启动错误，请输入需要处理的文件。exp: ./compareInvoices.exe Invoice.xlsx')

    while True:
        pass


