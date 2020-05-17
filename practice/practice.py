# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2020/03/26
# @Author  : 陈志鹏
# @File    : fetch.py

"""
提取一个工作簿的所有sheet信息， 合并成一个新的文件
"""

import os
import pandas as pd
import openpyxl
from PySide2.QtCore import *


class Practice(QObject):
    #message_singel = Signal(str)
    finish_singel = Signal()
    statusBar_singel = Signal(str)

    def __init__(self):
        super(Practice, self).__init__()

    def __del__(self):
        print('delete %s' % self.__class__.__name__)

    def practice1(self):
        df = pd.DataFrame({'ID':[1, 2, 3], 'Name':['Tim', 'Victor', 'Nick']})
        # df = pd.DataFrame()
        df.set_index('ID', inplace=True)
        file_path = os.getcwd()
        file_name = os.path.join(file_path, 'datas/practice1.xlsx')
        df.to_excel(file_name)
        print('Done!')

    def practice2(self):
        file_path = os.getcwd()
        file_name = os.path.join(file_path, 'datas/practice1.xlsx')
        df = pd.read_excel(file_name, header=None)
        df.columns = ['col1', 'col2']   # 添加列名
        print(df.head())
        print('Done!')

    def practice3(self):
        # fw = pd.DataFrame({'A': [1, 2]})
        #
        # fw1 = pd.DataFrame({'B': [3, 4]})
        #
        # fw2 = pd.DataFrame({'C': [5, 6]})
        #
        target_path = r'datas/practice3.xlsx'
        #
        # append_df_to_excel(target_path, fw, sheet_name='Sheet1', startcol=0, startrow=0, index=False, header=False)
        #
        # append_df_to_excel(target_path, fw1, sheet_name='Sheet1', startcol=1, startrow=0, index=False, header=False)
        #
        # append_df_to_excel(target_path, fw2, sheet_name='Sheet3', startcol=3, startrow=5, index=False, header=False)

        df = pd.read_excel(target_path, sheet_name='Sheet2', header=1)
        src_data = df['10月']
        print(src_data[:6])
        append_df_to_excel(target_path, src_data[:6], sheet_name='Sheet2', startcol=6, startrow=2, index=False,
                           header=False)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    # from openpyxl import load_workbook

    # import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = openpyxl.load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

if __name__ == '__main__':
    practice = Practice()
    # practice.practice1()
    # practice.practice2()
    practice.practice3()

